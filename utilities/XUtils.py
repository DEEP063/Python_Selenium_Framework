import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_column


def readData(file, sheetName, rownum, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columno).value


def writeData(file, sheetName, rownum, columno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columno).value = data
    workbook.save(file)

def fillGreenCoolor(file, sheetName, rownum, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenfill = PatternFill(start_color='60b212',end_color='60b212',fill_type='solid')
    sheet.cell(rownum,columno).fill = greenfill
    workbook.save(file)

def fillRedCoolor(file, sheetName, rownum, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redfill = PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')
    sheet.cell(rownum,columno).fill = redfill
    workbook.save(file)